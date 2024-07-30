from django.contrib.auth import authenticate
from django.shortcuts import render
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, generics
from rest_framework.exceptions import NotFound
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.authtoken.models import Token
import pandas
from api_gc.filters import *
from api_gc.models import *
from api_gc.serializers import *
from openpyxl import *
import os
import humanize
from django.conf import settings
from django.http import HttpResponse
from openpyxl.styles import *
from django.http import HttpResponse
from PyPDFForm import *
from reportlab.lib.pagesizes import *
from reportlab.pdfgen import canvas
from django.db.models import OuterRef, Subquery, DateField
from django.db.models.functions import Coalesce
from django.db.models import Count, Sum, Min, Max
# Create your views here.

class CreateUserView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class Login(APIView):
    permission_classes = []
    def post(self,request):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        app_name = 'api_sm'

        if user is not None:
            Token.objects.filter(user=user).delete()
            token, created = Token.objects.get_or_create(user=user)
            app_permissions = self.get_app_permissions(user, app_name)
            response = Response(status=status.HTTP_200_OK)
            role = '|'.join(list(app_permissions))
            response.set_cookie('token', token.key)
            response.set_cookie('role', role)
            return response
        else:
            return Response({'message': 'Informations d’identification non valides'}, status=status.HTTP_400_BAD_REQUEST)

    def get_app_permissions(self, user, app_name):
        # Get all permissions for the specified app
        all_permissions = user.get_all_permissions()
        app_permissions = set()

        for permission in all_permissions:
            if permission.split('.')[0] == app_name:
                app_permissions.add(permission)

        return app_permissions



class Whoami(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        return Response({'whoami':request.user.username}, status=status.HTTP_200_OK)

class Logout(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        Token.objects.get(user_id=request.user.id).delete()
        response=Response({'message': 'Vous etes déconnecté'}, status=status.HTTP_200_OK)
        response.delete_cookie('token')
        response.delete_cookie('role')
        return response

class ListImages(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Images.objects.all()
    serializer_class =ImagesSerilizer

class WhoamiView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        return Response({'whoami':request.user.username}, status=status.HTTP_200_OK)





class ListContract(generics.ListAPIView): # grouper par num contrat
    #permission_classes = [IsAuthenticated]
    queryset = Contrat.objects.all().order_by('-date_modification')
    serializer_class =ContratSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = ContratFilter


class ListPlaning(generics.ListAPIView):
    #permission_classes=[IsAuthenticated]
    queryset = Planing.objects.all().order_by('-date_modification')
    serializer_class = PlaningSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = PlaningFilter

class ImportPlaning(APIView):
    #permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        planing_file = request.FILES['file']
        contrat = Contrat_Latest.objects.get(id=request.data.get('contrat'))
        df=pandas.read_excel(planing_file, engine='openpyxl')
        sub_df = df.iloc[4:, 3:]
        sub_df.columns=sub_df.iloc[0]
        sub_df = sub_df[1:]
        sub_df = sub_df.reset_index(drop=True)
        sub_df.insert(0,' ',list(df.iloc[5:, 1]))
        sub_df.set_index(' ', inplace=True)
        for i, row in sub_df.iterrows():
            for col in sub_df.columns:
                code_prod=i
                date=col
                qte=row[col]
                
                dqe=DQECumule.objects.get(produit_id=code_prod)
                Planing(contrat=contrat,dqe=dqe,qte_livre=qte,date=date).save()

        return Response(status=status.HTTP_200_OK)
    

class PrintInv(APIView):
    def get(self,request):
        id=self.request.query_params.get('id',None)
        response={}
        factures=Factures.objects.get(id=id)
        details_qs=DetailFacture.objects.filter(facture=factures).order_by('detail__date_modification')
        cosider= InfoEntr.objects.all().first()
        details=[]
        response={
            'unite': f"({factures.id.split('_')[0]}) {Unite.objects.get(id=factures.id.split('_')[0]).libelle}",
            'client': factures.contrat.client.raison_social,
            'n_rc': factures.contrat.client.num_registre_commerce,
            'n_if':factures.contrat.client.nif,
            'ai':factures.contrat.client.article_imposition,
            'num_f':factures.id.split('_')[1],
            'date':factures.date.strftime('Le %d/%m/%Y'),
            "contrat":factures.contrat.numero,
            "rc":cosider.N_reg_c,
            'ref':cosider.reference,
            'rev':cosider.index,
            "nif":cosider.M_fiscale,
            "cap":f"{humanize.intcomma(round(cosider.Capital,2)).replace(',',' ')} DA",
            "tva": f"{factures.contrat.tva.id} % ",
            "rabais":f"{humanize.intcomma(round(factures.montant_rb,2)).replace(',',' ')} DA",
            "rg": f"{factures.contrat.rg} % ",
            'mht':f"{humanize.intcomma(round(factures.montant_facture_ht,2)).replace(',',' ')} DA",
            'mttc':f"{humanize.intcomma(round(factures.montant_facture_ttc,2)).replace(',',' ')} DA",
            'mrg':f"{humanize.intcomma(round(factures.montant_rg,2)).replace(',',' ')} DA",
            'cai':None,
        }

        
        for d in details_qs:
            
            obj={
                'bl':d.detail.num_bl,
                'ref_prod':d.detail.dqe.produit_id.id,
                'libelle':d.detail.dqe.produit_id.libelle,
                'UM':d.detail.dqe.produit_id.unite_m.id,
                'qte':d.detail.qte,
                'pu_ht':d.detail.dqe.prixproduit_id.prix_unitaire,
                't_ligne_ht':d.detail.montant
            }
            
            details.append(obj)
        
            
        response['details']=details
        return Response(response,status=status.HTTP_200_OK)
        



class PrintBL(APIView):
    def get(self,request):
    
        bl_id=self.request.query_params.get('bl',None)
        bl=BonLivraison.objects.get(id=bl_id)
       
        filled = FormWrapper(os.path.join(settings.MEDIA_ROOT, 'templates','Template_BL-Empty.pdf')).fill(
        {
            'unite':f"{bl.unite}",
            'num_bl':f"{bl.num_bl}",
            'date':f"{bl.date.strftime('Le %d/%m/%Y à %H:%M:%S')}",
            'code_client':f"{bl.client}",
            "rs_client":f"{bl.rs_client}",
            "conducteur":f"{bl.conducteur}",
            "camion":f"{bl.camion.matricule}",
            "tare":f"{bl.camion.tare}",
            "numero_permis_c":f"{bl.numero_permis_c}",
            "ptc":f"{bl.ptc}",
            "code_prod":f"{bl.dqe.produit_id}",
            "unite_m":f"{bl.dqe.produit_id.unite_m.id}",
            "qte":f"{bl.qte}",
            "pu":f"{bl.dqe.prixproduit_id.prix_unitaire}",
            "mht":f"{bl.montant}",
            "nap":f"{bl.net_a_payer}",
        },
        
        )
     


        # Return filled PDF as response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="filled_form.pdf"'
        response.write(filled.read())
        return response
        
        return None
        
    

class ListBL(generics.ListAPIView):
    # permission_classes = [IsAuthenticated]
    queryset = BonLivraison.objects.all().order_by('-date_modification')
    serializer_class = BonLivraisonSerializer
    filter_backends = [DjangoFilterBackend]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        mg = 0
        contrat=Contrat_Latest.objects.get(numero=self.request.query_params.get('contrat',None))
        response_data = super().list(request, *args, **kwargs).data
        date_overflow=None
        for q in queryset:
            mg += q.montant
            tmp=contrat.montant_ht-mg
          
        tmp=0
        for q in queryset:
            tmp += q.montant
            if(contrat.montant_ht-mg<0):
                date_overflow=q.date
                break
        if(date_overflow):
            date_overflow=date_overflow.strftime('%d-%m-%Y')
        else:
            date_overflow=None

        overflow=contrat.montant_ht-mg
        return Response({'bl': response_data,
                         'extra': {
                             'solde': contrat.montant_ht,
                             'mg': mg,
                             'ecart':overflow,
                             'date':date_overflow,

                         }}, status=status.HTTP_200_OK)



class AddBL(generics.CreateAPIView):
    # permission_classes = [IsAuthenticated]
    queryset= BonLivraison.objects.all()
    serializer_class = BonLivraisonSerializer
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            print(serializer)
            serializer.save()
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class AddEnc(generics.CreateAPIView):
    # permission_classes = [IsAuthenticated]
    queryset= Encaissement.objects.all()
    serializer_class = EncaissementSerializer

class GetEnc(generics.ListAPIView):
    # permission_classes = [IsAuthenticated]
    queryset= Encaissement.objects.all()
    serializer_class = EncaissementSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = EncFilter
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        mge = 0
        mgc  = 0
        
        for q in queryset:
            mge+=q.montant_encaisse
        
        latest_encaissement_subquery = queryset.filter(
            facture=OuterRef('facture')
        ).order_by('-date_encaissement').values('id')[:1]

        latest_encaissements = queryset.filter(
            id__in=Subquery(latest_encaissement_subquery)
        ).select_related('facture', 'mode_paiement')

        for le in latest_encaissements:
            mgc+=le.montant_creance

        response_data = super().list(request, *args, **kwargs).data
        return Response({
            'enc':response_data,
            'extra':{
                'mge': mge,
                'mgc':mgc,
                
            }
            
        }, status=status.HTTP_200_OK)



class PrintEnc(generics.ListAPIView):
    # permission_classes = [IsAuthenticated]
    queryset= Encaissement.objects.all().order_by('-facture')
    serializer_class = EncaissementSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = EncFilter
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        mge = 0
        mgc  = 0
        
        for q in queryset:
            mge+=q.montant_encaisse
        
        latest_encaissement_subquery = queryset.filter(
            facture=OuterRef('facture')
        ).order_by('-date_encaissement').values('id')[:1]

        latest_encaissements = queryset.filter(
            id__in=Subquery(latest_encaissement_subquery)
        ).select_related('facture', 'mode_paiement')

        for le in latest_encaissements:
            mgc+=le.montant_creance

        response_data = super().list(request, *args, **kwargs).data
        return Response({
            'enc':response_data,
            'extra':{
                'mge': mge,
                'mgc':mgc,
                
            }
            
        }, status=status.HTTP_200_OK)



class AddPlaning(generics.CreateAPIView):
    # permission_classes = [IsAuthenticated]
    queryset= Planing.objects.all()
    serializer_class = PlaningSerializer

class ListDQE(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = DQE.objects.all().order_by('-date_modification')
    serializer_class =DQESerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = DQEFilter
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        mt_ttc = 0
        mt_ht  = 0
        
        for q in queryset:
            mt_ht+=q.montant_qte
            mt_ttc+=q.montant_qte + (q.montant_qte* (q.contrat.tva.id/100))
        
        c=self.request.query_params.get('contrat__numero',None)
        a=self.request.query_params.get('contrat__avenant',None)
        contrat=Contrat.objects.get(numero=c,avenant=a)
        response_data = super().list(request, *args, **kwargs).data
        return Response({
            'dqe':response_data,
            'extra':{
                'ht': mt_ht,
                'ttc':mt_ttc,
                'tva': contrat.tva.id
            }
            
        }, status=status.HTTP_200_OK)



class ListDQECumulePlaning(APIView):
    def get(self,request):
        cid=self.request.query_params.get('contrat_id',None)
        if(cid):
            data=DQECumuleSerializer(DQECumule.objects.filter(contrat_id=cid),many=True).data
            contrat=Contrat.objects.get(id=cid)
        else:
            data=None
        
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

        if(data and contrat):
            template_path = os.path.join(settings.MEDIA_ROOT,'templates', 'planing.xlsx')
            wb = load_workbook(template_path)
            
            ws = wb.active  # Assuming single sheet for simplicity
            ws['C2']=contrat.numero
            ws['C3']=f'{contrat.avenant}'
            ws['C4']=f"{humanize.intcomma(round(contrat.montant_ht,2)).replace(',',' ')} DA"
            ws['C5']=f"{humanize.intcomma(round(contrat.montant_ttc,2)).replace(',',' ')} DA"





            # Write data to template
            row_offset = 7  # Start writing data from row 3 (assuming headers are in row 1 and 2)
            for row_index, item in enumerate(data, start=row_offset):
                for col_index, key in enumerate(['produit_id', 'libelle'], start=2):
                    cell = ws.cell(row=row_index, column=col_index)
                    cell.value = item[key]
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border 
            # Save modified workbook to response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=exported_data.xlsx'
            wb.save(response)
            return response
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)

class ListDQECumule(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = DQECumule.objects.all().order_by('-date_modification')
    serializer_class =DQECumuleSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = DQECumuleFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        mt_ttc = 0
        mt_ht  = 0
        c=self.request.query_params.get('code_contrat',None)
        contrat=Contrat_Latest.objects.get(numero=c)
        
        for q in queryset:
            mt_ht+=q.montant_qte
            mt_ttc+=q.montant_qte + (q.montant_qte* (contrat.tva.id/100))
        
        response_data = super().list(request, *args, **kwargs).data
        return Response({
            'dqe':response_data,
            'extra':{
                'ht': mt_ht,
                'ttc':mt_ttc,
                'tva': contrat.tva.id
            }
            
        }, status=status.HTTP_200_OK)


class ListClient(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Clients.objects.all().order_by('-date_modification')
    serializer_class =ClientSerilizer
    filter_backends = [DjangoFilterBackend]
    filterset_class= ClientFilter



class ListPrixProduit(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = PrixProduit.objects.all()
    serializer_class =PrixProduitSerializer
    filter_backends = [DjangoFilterBackend]



class AddClient(generics.CreateAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Clients.objects.all()
    serializer_class = ClientSerilizer
    

class AddDQE(generics.CreateAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = DQE.objects.all()
    serializer_class = DQESerializer


class GetUnites(APIView):
    def get(self,request):
        try:
            u=Unite.objects.all().values('id','libelle')
            return Response(u,status=status.HTTP_200_OK)
        except Unite.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)


class GetProds(APIView):
    def get(self,request):
        try:
            u=Produits.objects.all().values('id','libelle')
            return Response(u,status=status.HTTP_200_OK)
        except Produits.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)





class AddContrat(generics.CreateAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Contrat.objects.all()
    serializer_class = ContratSerializer






class AddCamion(generics.CreateAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Camion.objects.all()
    serializer_class = CamionSerializer


class ListCamion(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Camion.objects.all()
    serializer_class =CamionSerializer
    filter_backends = [DjangoFilterBackend]





class contratKeys(APIView):
    #permission_classes = [IsAuthenticated]
    def get(self,request):
        try:
            keys=Contrat.objects.all().values_list('numero', flat=True).distinct()
            result=[str(k) for k in keys]
            return Response(result,status=status.HTTP_200_OK)
        except Contrat.DoesNotExist:
            return Response({'message':'Pas de contrat'},status=status.HTTP_404_NOT_FOUND)


class DatesInv(APIView):
    #permission_classes = [IsAuthenticated]
    def get(self,request):
        try:
            contrat=self.request.query_params.get('cid',None)
            result = Factures.objects.filter(contrat__numero=contrat).aggregate(
                min_date=Min('date'),
                max_date=Max('date')
            )

            min_date = result['min_date']
            max_date = result['max_date']

            return Response({'min_date': min_date, 'max_date': max_date}, status=status.HTTP_200_OK)
        
        except Contrat.DoesNotExist:
            return Response({'message':'Pas de Facture(s)'},status=status.HTTP_404_NOT_FOUND)





class AvenantKeys(APIView):
    #permission_classes = [IsAuthenticated]
    def get(self,request):
        num=self.request.query_params.get('num',None)
        try:
            keys=Contrat.objects.filter(numero=num).values_list('avenant', flat=True).distinct()
            result=[str(k) for k in keys]
            return Response(result,status=status.HTTP_200_OK)
        except Contrat.DoesNotExist:
            return Response({'message':'Pas d\' avenant'},status=status.HTTP_404_NOT_FOUND)


class LastAvenantKeys(APIView):
    #permission_classes = [IsAuthenticated]
    def get(self,request):
        num=self.request.query_params.get('num',None)
        try:
            key=Contrat.objects.filter(numero=num).latest('avenant').avenant
            
            return Response(key,status=status.HTTP_200_OK)
        except Contrat.DoesNotExist:
            return Response({'message':'Pas d\' avenant'},status=status.HTTP_404_NOT_FOUND)




class AddFacture(generics.CreateAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Factures.objects.all()
    serializer_class = FactureSerializer




class UpdateDQE(generics.UpdateAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = DQE.objects.all()
    serializer_class = DQESerializer
    lookup_field = "pk"
    def get_object(self):
        pk = self.request.data.get(DQE._meta.pk.name)
        try:
            obj = DQE.objects.get(pk=pk)
        except DQE.DoesNotExist:
            raise NotFound("Object n'éxiste pas")
        self.check_object_permissions(self.request, obj)

        return obj


class DeleteDQE(generics.DestroyAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = DQE.objects.all()
    serializer_class = DQESerializer

    def delete(self, request, *args, **kwargs):
        pk = request.data.get(DQE._meta.pk.name)
        if pk:
            queryset = self.filter_queryset(self.get_queryset())
            queryset = queryset.filter(pk__in=pk)
            self.perform_destroy(queryset)

        return Response({'Message': pk}, status=status.HTTP_200_OK)



class DeleteBL(generics.DestroyAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = BonLivraison.objects.all()
    serializer_class = BonLivraisonSerializer

    def delete(self, request, *args, **kwargs):
        pks = request.data.get('id')
        if pks:
            queryset = self.filter_queryset(self.get_queryset())
            queryset = queryset.filter(pk__in=pks)
            for qs in queryset:
                qs.delete()

        return Response({'Message': pks}, status=status.HTTP_200_OK)


class DeleteInvoice(generics.DestroyAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Factures.objects.all()
    serializer_class = FactureSerializer
     
    def delete(self, request, *args, **kwargs):
        pks = self.request.data.get('id', [])
        
        if pks:
            queryset = self.filter_queryset(self.get_queryset())
            queryset = queryset.filter(pk__in=pks)
            for q in queryset:
                    q.delete()
                
        return Response({'Message': pks}, status=status.HTTP_200_OK)











class ListFacture(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Factures.objects.all().order_by('-date_modification')
    serializer_class =FactureSerializer
    filter_backends = [DjangoFilterBackend]


    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        rg_total = 0
        creance = 0
        response_data = super().list(request, *args, **kwargs).data
        for q in queryset:
            rg_total += q.montant_rg




        return Response({'facture': response_data,
                         'extra': {

                             'rg_total': rg_total,
                             'creance': creance,

                         }}, status=status.HTTP_200_OK)





class ListDetail(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = DetailFacture.objects.all().order_by('-date_modification')
    serializer_class =DetailFactureSerializer
    filter_backends = [DjangoFilterBackend]




class ListAvance(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    queryset = Avances.objects.all().order_by('-date_modification')
    serializer_class =AvanceSerializer
    filter_backends = [DjangoFilterBackend]


    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        total = 0
        rst = 0
        response_data = super().list(request, *args, **kwargs).data
        for q in queryset:
            total = total + q.montant_avance
            rst = rst + q.montant_restant

        return Response({'avances': response_data,
                         'extra': {

                             'total': total,
                             'rst': rst,


                         }}, status=status.HTTP_200_OK)





class AddAvance(generics.CreateAPIView):
    # permission_classes = [IsAuthenticated]
    queryset= Avances.objects.all()
    serializer_class = AvanceSerializer